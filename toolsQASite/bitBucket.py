import unittest
from selenium import webdriver
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


class Bitbucket(unittest.TestCase):

    def setUp(self):
        self.driver = webdriver.Chrome()
        folderPath = os.getcwd()
        self.filePath=os.path.join(folderPath, "data", "openpyxl.xlsx")

    def testAllRights(self):
        driver = self.driver
        book=load_workbook(self.filePath)
        sheet = book.get_sheet_by_name("logIn")

        email=sheet['B2']
        password=sheet['C2']
        driver.get("https://id.atlassian.com/login")
        time.sleep(5)
        mailEL = driver.find_element_by_id("username")
        mailEL.clear()
        mailEL.send_keys(email.value)
        ContEl = driver.find_element_by_id("login-submit")
        time.sleep(3)
        ContEl.click()
        print("Klik na ContEL")
        passEL = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "password")))
        print("nasao passEL")
        time.sleep(3)
        passEL.clear()
        passEL.send_keys(password.value)
        time.sleep(3)
        ContEl.submit()
        time.sleep(3)
        title = driver.title
        assert title=="Atlassian account"
        sheet.cell(row=2, column=4).value = "PASS"
        book.save(self.filePath)
        time.sleep(10)


    def tearDown(self):
        self.driver.close()

if __name__ == "__main__":
    unittest.main()
