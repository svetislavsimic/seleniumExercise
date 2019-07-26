import unittest
from selenium import webdriver
import os
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains

class Fastidious(unittest.TestCase):

    def setUp(self):
        self.driver = webdriver.Chrome()
        working_folder = os.getcwd()
        data_folder = os.path.dirname(working_folder)
        print(data_folder)
        self.filePath = os.path.join(data_folder, "data", "fastidious.xlsx")

    def testLogInAdmin(self):
        driver = self.driver
        book = load_workbook(self.filePath)
        sheet = book["login"]
        username = sheet['B2'].value
        password = sheet['C2'].value
        driver.get("https://fastidious-app.netlify.com")
        wait = WebDriverWait(driver, 10)
        assert "Fastidious Admin" in driver.title

        user_field = driver.find_element_by_xpath("//input[contains(@name,'email')]")
        password_field = driver.find_element_by_xpath("//input[contains(@name,'password')]")
        login_btn = driver.find_element_by_xpath("//button[@type='submit']")
        logged_in_url = r"https://fastidious-app.netlify.com/dashboard"

        user_field.send_keys(username)
        password_field.send_keys(password)
        login_btn.click()

        wait.until(EC.url_contains(logged_in_url))
        assert driver.current_url == logged_in_url

        sheet.cell(row=2, column=4).value = "PASS"
        book.save(self.filePath)

    def testBlankLogIn(self):
        driver = self.driver
        book = load_workbook(self.filePath)
        sheet = book["login"]
        driver.get("https://fastidious-app.netlify.com")
        driver.maximize_window()
        assert "Fastidious Admin" in driver.title

        user_field = driver.find_element_by_xpath("//input[contains(@name,'email')]")
        password_field = driver.find_element_by_xpath("//input[contains(@name,'password')]")
        login_btn = driver.find_element_by_xpath("//button[@type='submit']")

        user_field.send_keys()
        password_field.send_keys()
        login_btn.click()

        email_alert = driver.find_element_by_xpath("//div[contains(text(),'Email is required')]")
        password_alert = driver.find_element_by_xpath("//div[contains(text(),'Password is required')]")

        assert email_alert, password_alert in driver.page_source

        sheet.cell(row=3, column=4).value = "PASS"
        book.save(self.filePath)

    def testOnlyUsername(self):
        driver = self.driver
        book = load_workbook(self.filePath)
        sheet = book["login"]
        username = sheet['B2'].value
        driver.get("https://fastidious-app.netlify.com")
        driver.maximize_window()
        assert "Fastidious Admin" in driver.title

        user_field = driver.find_element_by_xpath("//input[contains(@name,'email')]")
        password_field = driver.find_element_by_xpath("//input[contains(@name,'password')]")
        login_btn = driver.find_element_by_xpath("//button[@type='submit']")

        user_field.send_keys(username)
        password_field.send_keys()
        login_btn.click()

        password_alert = driver.find_element_by_xpath("//div[contains(text(),'Password is required')]")
        password_alert.is_displayed()
        assert 'Password is required' in driver.page_source

        sheet.cell(row=4, column=4).value = "PASS"
        book.save(self.filePath)

    def testOnlyPassword(self):
        driver = self.driver
        book = load_workbook(self.filePath)
        sheet = book["login"]
        password = sheet['C2'].value
        driver.get("https://fastidious-app.netlify.com")
        driver.maximize_window()
        assert "Fastidious Admin" in driver.title

        user_field = driver.find_element_by_xpath("//input[contains(@name,'email')]")
        password_field = driver.find_element_by_xpath("//input[contains(@name,'password')]")
        login_btn = driver.find_element_by_xpath("//button[@type='submit']")

        password_field.send_keys(password)
        user_field.send_keys()
        login_btn.click()

        driver.find_element_by_xpath("//div[contains(text(),'Email is required')]")
        assert 'Email is required' in driver.page_source

        sheet.cell(row=5, column=4).value = "PASS"
        book.save(self.filePath)

    def testWrongUsername(self):
        driver = self.driver
        book = load_workbook(self.filePath)
        sheet = book["login"]
        username = sheet['B6'].value
        password = sheet['C6'].value
        driver.get("https://fastidious-app.netlify.com")
        driver.maximize_window()
        assert "Fastidious Admin" in driver.title
        wait = WebDriverWait(driver, 20)

        user_field = driver.find_element_by_xpath("//input[contains(@name,'email')]")
        password_field = driver.find_element_by_xpath("//input[contains(@name,'password')]")
        login_btn = driver.find_element_by_xpath("//button[@type='submit']")

        user_field.send_keys(username)
        password_field.send_keys(password)
        login_btn.click()

        wait.until(EC.presence_of_element_located((
            By.XPATH, "//div[@class='message'][contains(.,'Unable to login. Try again!')]"
        )))
        assert 'Unable to login. Try again!' in driver.page_source
        sheet.cell(row=6, column=4).value = "PASS"
        book.save(self.filePath)

    def testWrongPassword(self):
        driver = self.driver
        book = load_workbook(self.filePath)
        sheet = book["login"]
        username = sheet['B7'].value
        password = sheet['C7'].value
        driver.get("https://fastidious-app.netlify.com")
        driver.maximize_window()
        assert "Fastidious Admin" in driver.title
        wait = WebDriverWait(driver, 20)

        user_field = driver.find_element_by_xpath("//input[contains(@name,'email')]")
        password_field = driver.find_element_by_xpath("//input[contains(@name,'password')]")
        login_btn = driver.find_element_by_xpath("//button[@type='submit']")

        user_field.send_keys(username)
        password_field.send_keys(password)
        login_btn.click()

        wait.until(EC.presence_of_element_located((
            By.XPATH, "//div[@class='message'][contains(.,'Unable to login. Try again!')]"
        )))
        assert 'Unable to login. Try again!' in driver.page_source
        sheet.cell(row=7, column=4).value = "PASS"
        book.save(self.filePath)

    def testWrongUserAndPass(self):
        driver = self.driver
        book = load_workbook(self.filePath)
        sheet = book["login"]
        username = sheet['B8'].value
        password = sheet['C8'].value
        driver.get("https://fastidious-app.netlify.com")
        driver.maximize_window()
        assert "Fastidious Admin" in driver.title
        wait = WebDriverWait(driver, 20)

        user_field = driver.find_element_by_xpath("//input[contains(@name,'email')]")
        password_field = driver.find_element_by_xpath("//input[contains(@name,'password')]")
        login_btn = driver.find_element_by_xpath("//button[@type='submit']")

        user_field.send_keys(username)
        password_field.send_keys(password)
        login_btn.click()

        wait.until(EC.presence_of_element_located((
            By.XPATH, "//div[@class='message'][contains(.,'Unable to login. Try again!')]"
        )))
        assert 'Unable to login. Try again!' in driver.page_source
        sheet.cell(row=8, column=4).value = "PASS"
        book.save(self.filePath)

    def testInOutInOut(self):
        driver = self.driver
        book = load_workbook(self.filePath)
        sheet = book["login"]
        driver.get("https://fastidious-app.netlify.com")
        driver.maximize_window()
        action = ActionChains(driver)
        assert "Fastidious Admin" in driver.title
        wait = WebDriverWait(driver, 20)
        logged_in_url = r"https://fastidious-app.netlify.com/dashboard"

        user_field = driver.find_element_by_xpath("//input[contains(@name,'email')]")
        password_field = driver.find_element_by_xpath("//input[contains(@name,'password')]")
        login_btn = driver.find_element_by_xpath("//button[@type='submit']")

        user_field.send_keys(sheet['B2'].value)  # user ok
        password_field.send_keys(sheet['C7'].value)  # pass nok
        login_btn.click()

        wait.until(EC.visibility_of_element_located((
            By.XPATH, "//div[@class='message'][contains(.,'Unable to login. Try again!')]"
        )))
        assert 'Unable to login. Try again!' in driver.page_source

        password_field.send_keys(Keys.CONTROL + "a")
        password_field.send_keys(Keys.DELETE)
        password_field.send_keys(sheet['C6'].value)  # pass ok
        action.send_keys(Keys.ENTER).perform()

        wait.until(EC.url_to_be(logged_in_url))
        assert driver.current_url == logged_in_url
        driver.find_element_by_xpath("//i[@class='fas fa-sign-out-alt']").click()  # go back to login

        wait.until(EC.url_to_be(r"https://fastidious-app.netlify.com/login"))
        assert r"https://fastidious-app.netlify.com/login" in driver.current_url

        # log in again
        driver.find_element_by_name("email").send_keys(sheet['B6'].value)  # user nok
        driver.find_element_by_name("password").send_keys(sheet['C2'].value)  # pass ok
        action.send_keys(Keys.ENTER).perform()

        wait.until(EC.visibility_of_element_located((
            By.XPATH, "//div[@class='message'][contains(.,'Unable to login. Try again!')]"
        )))
        assert 'Unable to login. Try again!' in driver.page_source

        username = driver.find_element_by_name("email")
        username.send_keys(Keys.CONTROL + "a")
        username.send_keys(Keys.DELETE)
        username.send_keys(sheet['B2'].value) # user ok
        action.send_keys(Keys.ENTER).perform()

        wait.until(EC.url_to_be(logged_in_url)) # logged in
        assert driver.current_url == logged_in_url
        driver.find_element_by_xpath("//i[@class='fas fa-sign-out-alt']").click()

        wait.until(EC.url_to_be(r"https://fastidious-app.netlify.com/login"))
        assert r"https://fastidious-app.netlify.com/login" in driver.current_url

        sheet.cell(row=9, column=4).value = "PASS"
        book.save(self.filePath)

    def tearDown(self):
        self.driver.close()


if __name__ == "__main__":
    unittest.main()
